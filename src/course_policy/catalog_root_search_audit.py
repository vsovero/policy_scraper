"""Audit catalog/archive root discovery for the current Phase 3 test institutions."""

from __future__ import annotations

import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_config import repo_root_from_cwd
from .batch2_root_check import candidate_urls_for_task, likely_catalog_root, root_priority
from .batch2_year_candidates import candidate_archive_urls
from .batch3_discovery import (
    build_legacy_leads,
    build_year_candidates,
    source_root_tasks,
)
from .catalog_retrieval import retrieve_url


DATA_DIR = Path("../data_policy_pipeline")
INTERIM_DIR = DATA_DIR / "interim"
REVIEW_DIR = DATA_DIR / "review"
LOG_DIR = DATA_DIR / "logs"

INSTITUTION_UNIVERSE_INPUT = INTERIM_DIR / "institution_universe.csv"
LEGACY_EVIDENCE_LINKS_INPUT = INTERIM_DIR / "legacy_evidence_links.csv"
SPOTCHECK_WORKBOOK_INPUT = REVIEW_DIR / "catalog_url_spotcheck_mockup.xlsx"

ROOT_SEARCH_AUDIT_OUTPUT = REVIEW_DIR / "catalog_root_search_audit.xlsx"
ROOT_SEARCH_SUMMARY_OUTPUT = LOG_DIR / "catalog_root_search_audit_summary.md"

RETRIEVED_STATUSES = {"retrieved", "retrieved_truncated"}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def current_test_institutions(repo_root: Path) -> pd.DataFrame:
    spotcheck_path = repo_root / SPOTCHECK_WORKBOOK_INPUT
    if spotcheck_path.exists():
        spotcheck = pd.read_excel(spotcheck_path, sheet_name="spotcheck_mockup")
        unitids = sorted(spotcheck["unitid"].dropna().astype(int).unique())
    else:
        unitids = []
    universe = pd.read_csv(repo_root / INSTITUTION_UNIVERSE_INPUT, low_memory=False)
    if unitids:
        universe = universe.loc[universe["unitid"].astype(int).isin(unitids)].copy()
    return universe[["unitid", "institution_name", "state", "webaddr"]].drop_duplicates().sort_values("institution_name")


def root_candidates_for_test_set(
    repo_root: Path,
    *,
    timeout_seconds: int,
    max_candidates_per_institution: int,
    workers: int,
) -> pd.DataFrame:
    institutions = current_test_institutions(repo_root)
    institutions = institutions.copy()
    institutions["batch3_rank"] = range(1, len(institutions) + 1)
    institutions["pilot_rank"] = institutions["batch3_rank"]
    institutions["pilot_case_types"] = "root_search_audit"
    legacy_links = pd.read_csv(repo_root / LEGACY_EVIDENCE_LINKS_INPUT, low_memory=False)
    legacy_leads = build_legacy_leads(institutions, legacy_links)
    tasks = source_root_tasks(institutions, legacy_leads)

    jobs = []
    for _, task in tasks.sort_values(["batch3_rank", "unitid"]).iterrows():
        candidates = sorted(candidate_urls_for_task(task, legacy_leads), key=prefetch_priority)[:max_candidates_per_institution]
        for idx, candidate in enumerate(candidates, 1):
            jobs.append((task, idx, candidate))

    rows = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(fetch_root_candidate, task, idx, candidate, timeout_seconds): (task, idx, candidate)
            for task, idx, candidate in jobs
        }
        for future in as_completed(future_map):
            rows.append(future.result())
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(["batch3_rank", "root_priority", "candidate_rank"])


def fetch_root_candidate(task: pd.Series, idx: int, candidate: dict[str, str], timeout_seconds: int) -> dict[str, object]:
    result = retrieve_url(
        candidate["candidate_url"],
        timeout_seconds=timeout_seconds,
        max_bytes=1_000_000,
    )
    is_likely = likely_catalog_root(result, candidate["candidate_url"], candidate["candidate_source_type"])
    return {
        "batch3_rank": int(task["batch3_rank"]),
        "unitid": int(task["unitid"]),
        "institution_name": task["institution_name"],
        "candidate_rank": idx,
        "candidate_url": candidate["candidate_url"],
        "candidate_source_type": candidate["candidate_source_type"],
        "retrieval_status": result["retrieval_status"],
        "http_status": result["http_status"],
        "final_url": result["final_url"],
        "content_type": result["content_type"],
        "page_title": result["page_title"],
        "year_hints": result["year_hints"],
        "link_count": len(result.get("link_records", [])),
        "likely_catalog_root": is_likely,
        "root_priority": root_priority(candidate["candidate_source_type"], result, candidate["candidate_url"]),
        "created_at": utc_now(),
    }


def prefetch_priority(candidate: dict[str, str]) -> tuple[int, str]:
    source_type = candidate["candidate_source_type"]
    url = candidate["candidate_url"].lower()
    if source_type in {"legacy_derived_archive_root", "legacy_derived_repository_collection"}:
        return (1, url)
    if source_type.startswith("generated") and "catalogarchive" in url:
        return (2, url)
    if source_type in {"generated_catalog_subdomain", "generated_catalogs_subdomain"}:
        return (3, url)
    if source_type == "generated_catalog_archive_path":
        return (4, url)
    if source_type.startswith("generated"):
        return (5, url)
    if source_type == "legacy_parent_url":
        return (6, url)
    if source_type == "legacy_url":
        return (7, url)
    return (9, url)


def preferred_roots(root_candidates: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (rank, unitid, name), group in root_candidates.groupby(["batch3_rank", "unitid", "institution_name"], dropna=False):
        usable = group.loc[group["likely_catalog_root"].fillna(False)].copy()
        if usable.empty:
            rows.append(
                {
                    "batch3_rank": rank,
                    "unitid": unitid,
                    "institution_name": name,
                    "decision_status": "source_root_not_found",
                    "preferred_source_root_url": "",
                    "preferred_source_root_type": "",
                    "preferred_source_root_title": "",
                    "root_retrieval_status": "",
                    "root_year_hints": "",
                }
            )
            continue
        preferred = usable.sort_values(["root_priority", "candidate_rank"]).iloc[0]
        rows.append(
            {
                "batch3_rank": rank,
                "unitid": unitid,
                "institution_name": name,
                "decision_status": "preferred_source_root_identified",
                "preferred_source_root_url": preferred["candidate_url"],
                "preferred_source_root_type": preferred["candidate_source_type"],
                "preferred_source_root_title": preferred["page_title"],
                "root_retrieval_status": preferred["retrieval_status"],
                "root_year_hints": preferred["year_hints"],
            }
        )
    return pd.DataFrame(rows).sort_values(["batch3_rank", "unitid"])


def year_candidate_probe(preferred: pd.DataFrame, *, timeout_seconds: int, workers: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    archive_rows = []
    result_by_url: dict[str, dict[str, object]] = {}
    archive_jobs = []
    preferred_rows = list(preferred.loc[preferred["decision_status"].eq("preferred_source_root_identified")].to_dict("records"))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(fetch_archive_probe_url, clean_text(row["preferred_source_root_url"]), timeout_seconds): row
            for row in preferred_rows
        }
        for future in as_completed(future_map):
            row = future_map[future]
            root_url, result = future.result()
            result_by_url[root_url] = result
            for archive in candidate_archive_urls(result, root_url):
                archive_jobs.append((row, archive))

    missing_archive_jobs = []
    for row, archive in archive_jobs:
        archive_url = archive["archive_url"]
        if archive_url not in result_by_url:
            missing_archive_jobs.append((row, archive))

    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(fetch_archive_probe_url, archive["archive_url"], timeout_seconds): (row, archive)
            for row, archive in missing_archive_jobs
        }
        for future in as_completed(future_map):
            archive_url, result = future.result()
            result_by_url[archive_url] = result

    for row, archive in archive_jobs:
        archive_url = archive["archive_url"]
        archive_result = result_by_url[archive_url]
        archive_rows.append(
            {
                "batch3_rank": int(row["batch3_rank"]),
                "unitid": int(row["unitid"]),
                "institution_name": row["institution_name"],
                "preferred_source_root_url": clean_text(row["preferred_source_root_url"]),
                "archive_url": archive_url,
                "archive_source": archive["archive_source"],
                "archive_link_text": archive["archive_link_text"],
                "retrieval_status": archive_result["retrieval_status"],
                "http_status": archive_result["http_status"],
                "final_url": archive_result["final_url"],
                "content_type": archive_result["content_type"],
                "page_title": archive_result["page_title"],
                "year_hints": archive_result["year_hints"],
                "link_count": len(archive_result.get("link_records", [])),
                "local_source_path": "",
                "created_at": utc_now(),
            }
        )
    archive_pages = pd.DataFrame(archive_rows)
    if archive_pages.empty:
        return archive_pages, pd.DataFrame()
    year_candidates = build_year_candidates(archive_pages, result_by_url)
    return archive_pages, year_candidates


def fetch_archive_probe_url(url: str, timeout_seconds: int) -> tuple[str, dict[str, object]]:
    request_timeout = max(timeout_seconds, 20) if "/digital/api/search/" in url else timeout_seconds
    return url, retrieve_url(url, timeout_seconds=request_timeout, max_bytes=5_000_000)


def institution_summary(preferred: pd.DataFrame, year_candidates: pd.DataFrame) -> pd.DataFrame:
    if year_candidates.empty:
        coverage = pd.DataFrame(columns=["unitid", "candidate_year_count", "candidate_start_year", "candidate_end_year"])
    else:
        coverage = (
            year_candidates.groupby("unitid", as_index=False)
            .agg(
                candidate_year_count=("target_year", "nunique"),
                candidate_start_year=("target_year", "min"),
                candidate_end_year=("target_year", "max"),
                candidate_url_count=("candidate_url", "nunique"),
            )
        )
    out = preferred.merge(coverage, on="unitid", how="left")
    for col in ["candidate_year_count", "candidate_start_year", "candidate_end_year", "candidate_url_count"]:
        out[col] = out[col].fillna(0).astype(int)
    out["root_search_status"] = "needs_followup"
    out.loc[out["candidate_year_count"].ge(15), "root_search_status"] = "strong_root_found"
    out.loc[out["candidate_year_count"].between(1, 14), "root_search_status"] = "partial_root_found"
    out.loc[out["decision_status"].eq("source_root_not_found"), "root_search_status"] = "root_not_found"
    out["recommended_next_step"] = out["root_search_status"].map(
        {
            "strong_root_found": "Use this root for year-candidate expansion and retrieval.",
            "partial_root_found": "Inspect root pagination/search and sibling archive links before moving on.",
            "needs_followup": "Run targeted web/archive search or source-specific parser.",
            "root_not_found": "Run targeted web/archive search.",
        }
    )
    return out.sort_values(["institution_name"])


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in ws.columns:
            max_len = max(len(clean_text(cell.value)) for cell in column_cells[:200])
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 70)
    wb.save(path)


def write_outputs(
    repo_root: Path,
    summary: pd.DataFrame,
    root_candidates: pd.DataFrame,
    archive_pages: pd.DataFrame,
    year_candidates: pd.DataFrame,
) -> tuple[Path, Path]:
    workbook_path = repo_root / ROOT_SEARCH_AUDIT_OUTPUT
    summary_path = repo_root / ROOT_SEARCH_SUMMARY_OUTPUT
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="institution_summary", index=False)
        root_candidates.to_excel(writer, sheet_name="root_candidates", index=False)
        archive_pages.to_excel(writer, sheet_name="archive_probe", index=False)
        year_candidates.to_excel(writer, sheet_name="year_candidates_probe", index=False)
    format_workbook(workbook_path)

    lines = [
        "# Catalog Root Search Audit",
        "",
        f"Generated at: {utc_now()}",
        "",
        f"Workbook: `{workbook_path}`",
        "",
        "## Root Search Status",
        "",
    ]
    for status, count in summary["root_search_status"].value_counts(dropna=False).items():
        lines.append(f"- {status}: {count}")
    summary_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return workbook_path, summary_path


def run(repo_root: Path, *, timeout_seconds: int, max_candidates_per_institution: int, workers: int) -> tuple[Path, Path]:
    repo_root = repo_root.resolve()
    roots = root_candidates_for_test_set(
        repo_root,
        timeout_seconds=timeout_seconds,
        max_candidates_per_institution=max_candidates_per_institution,
        workers=workers,
    )
    preferred = preferred_roots(roots)
    archive_pages, year_candidates = year_candidate_probe(preferred, timeout_seconds=timeout_seconds, workers=workers)
    summary = institution_summary(preferred, year_candidates)
    return write_outputs(repo_root, summary, roots, archive_pages, year_candidates)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=None)
    parser.add_argument("--timeout-seconds", type=int, default=8)
    parser.add_argument("--max-candidates-per-institution", type=int, default=24)
    parser.add_argument("--workers", type=int, default=8)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    repo_root = args.repo_root or repo_root_from_cwd()
    workbook_path, summary_path = run(
        repo_root,
        timeout_seconds=args.timeout_seconds,
        max_candidates_per_institution=args.max_candidates_per_institution,
        workers=args.workers,
    )
    print(f"workbook: {workbook_path}")
    print(f"summary: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
