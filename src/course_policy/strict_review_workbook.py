"""Create a focused review workbook for the strict Phase 3 catalog pilot."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_config import repo_root_from_cwd


DATA_DIR = Path("../data_policy_pipeline")
INTERIM_DIR = DATA_DIR / "interim"
REVIEW_DIR = DATA_DIR / "review"

STRICT_YEAR_COVERAGE_INPUT = INTERIM_DIR / "catalog_year_coverage_strict_pilot.csv"
STRICT_RETRIEVAL_COVERAGE_INPUT = INTERIM_DIR / "catalog_retrieval_coverage_strict_pilot.csv"
STRICT_INVENTORY_INPUT = INTERIM_DIR / "catalog_inventory_strict_pilot.csv"
PANEL_CANDIDATES_INPUT = INTERIM_DIR / "catalog_panel_candidates_strict_pilot.csv"
PANEL_YEAR_STATUS_INPUT = INTERIM_DIR / "catalog_panel_year_status_strict_pilot.csv"
PANEL_RETRIEVAL_COVERAGE_INPUT = INTERIM_DIR / "catalog_panel_retrieval_coverage_strict_pilot.csv"
PANEL_RETRIEVED_YEAR_COVERAGE_INPUT = INTERIM_DIR / "catalog_panel_year_coverage_retrieved_strict_pilot.csv"
SOURCE_ROOT_PLAN_INPUT = INTERIM_DIR / "catalog_source_root_plan_strict_pilot.csv"
ESCALATION_QUEUE_INPUT = INTERIM_DIR / "catalog_first_pass_escalation_queue_strict_pilot.csv"
CURRENT_SOURCE_TRACE_INPUT = INTERIM_DIR / "catalog_current_process_source_trace_strict_pilot.csv"
CURRENT_YEAR_TRACE_INPUT = INTERIM_DIR / "catalog_current_process_year_trace_strict_pilot.csv"
OCR_VISUAL_CONFIRMATION_INPUT = INTERIM_DIR / "catalog_ocr_visual_confirmation_strict_pilot.csv"

STRICT_REVIEW_WORKBOOK_OUTPUT = REVIEW_DIR / "strict_catalog_pilot_review.xlsx"

SUMMARY_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "state",
    "strict_pilot_reason",
    "institution_years",
    "covered_years",
    "missing_years",
    "coverage_rate",
    "legacy_evidence_years",
    "source_rows",
    "strict_source_rows",
    "review_source_rows",
]

YEAR_COVERAGE_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "target_year",
    "has_strict_catalog_source",
    "source_status",
    "source_id",
    "catalog_year_start",
    "catalog_year_end",
    "catalog_year_evidence_type",
    "catalog_year_evidence_text",
    "candidate_url",
    "retrieval_method",
    "local_source_path",
    "prior_evidence_status",
    "legacy_evidence_row_count",
    "legacy_needs_review",
    "needs_human_review",
    "review_reason",
]

SOURCE_EVIDENCE_COLUMNS = [
    "source_id",
    "unitid",
    "institution_name",
    "target_year",
    "source_retrieved",
    "strict_covers_target_year",
    "catalog_year_evidence_type",
    "catalog_year_start",
    "catalog_year_end",
    "catalog_year_evidence_text",
    "candidate_url",
    "best_final_url",
    "best_attempt_method",
    "best_content_type",
    "best_page_title",
    "best_year_hints",
    "local_source_path",
    "sha256",
    "legacy_workbook",
    "legacy_sheet_name",
    "legacy_excel_row",
    "legacy_selected_as_prior_evidence",
    "legacy_needs_review",
    "legacy_review_reasons",
    "strict_coverage_reason",
]

NEEDS_REVIEW_COLUMNS = [
    "review_type",
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "target_year",
    "source_id",
    "review_reason",
    "candidate_url",
    "best_final_url",
    "catalog_year_evidence_type",
    "catalog_year_start",
    "catalog_year_end",
    "catalog_year_evidence_text",
    "local_source_path",
    "legacy_excel_row",
    "legacy_review_reasons",
]

PANEL_YEAR_STATUS_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "target_year",
    "has_strict_catalog_source",
    "current_strict_source_id",
    "candidate_status",
    "candidate_source_id",
    "candidate_title",
    "candidate_url",
    "candidate_review_reason",
    "strict_pilot_reason",
]

PANEL_CANDIDATE_COLUMNS = [
    "source_id",
    "unitid",
    "institution_name",
    "candidate_url",
    "source_title",
    "catalog_year_start",
    "catalog_year_end",
    "discovery_method",
    "source_kind",
    "source_status",
    "archive_page_url",
    "needs_human_review",
    "review_reason",
    "created_at",
]

PANEL_RETRIEVAL_COLUMNS = [
    "source_id",
    "pilot_rank",
    "unitid",
    "institution_name",
    "target_year",
    "candidate_url",
    "source_retrieved",
    "strict_covers_target_year",
    "catalog_year_evidence_type",
    "catalog_year_start",
    "catalog_year_end",
    "catalog_year_evidence_text",
    "best_retrieval_status",
    "best_attempt_method",
    "best_final_url",
    "best_content_type",
    "best_page_title",
    "local_source_path",
    "sha256",
    "strict_coverage_reason",
]

PANEL_RETRIEVED_YEAR_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "target_year",
    "has_strict_catalog_source",
    "source_status",
    "source_id",
    "catalog_year_start",
    "catalog_year_end",
    "catalog_year_evidence_type",
    "catalog_year_evidence_text",
    "candidate_url",
    "retrieval_method",
    "local_source_path",
    "prior_evidence_status",
    "legacy_evidence_row_count",
    "legacy_needs_review",
    "needs_human_review",
    "review_reason",
]

SOURCE_ROOT_PLAN_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "source_root_role",
    "source_root_name",
    "source_root_url",
    "source_root_type",
    "root_scope",
    "first_pass_decision",
    "first_ay_observed",
    "last_ay_observed",
    "archive_bound_basis",
    "fallback_order",
    "notes",
    "created_at",
]

ESCALATION_QUEUE_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "escalation_bucket",
    "source_root_name",
    "source_root_url",
    "reason",
    "recommended_next_step",
    "created_at",
]

CURRENT_SOURCE_TRACE_COLUMNS = [
    "source_id",
    "unitid",
    "institution_name",
    "target_year",
    "candidate_url",
    "source_retrieved",
    "strict_covers_target_year",
    "catalog_year_start",
    "catalog_year_end",
    "catalog_year_evidence_type",
    "best_attempt_method",
    "source_status",
    "actual_process_role",
    "source_trace_origin",
    "review_reason",
    "local_source_path",
    "created_at",
]

CURRENT_YEAR_TRACE_COLUMNS = [
    "strict_pilot_rank",
    "unitid",
    "institution_name",
    "target_year",
    "has_strict_catalog_source",
    "source_id",
    "candidate_status",
    "candidate_source_id",
    "candidate_title",
    "candidate_url",
    "actual_process_role",
    "reverse_engineered_step",
    "review_reason",
    "candidate_review_reason",
    "created_at",
]

OCR_VISUAL_COLUMNS = [
    "source_id",
    "unitid",
    "institution_name",
    "source_title",
    "catalog_year_start",
    "catalog_year_end",
    "candidate_url",
    "retrieval_status",
    "pdf_recovery_method",
    "pdf_attempt_url",
    "render_status",
    "confirmation_status",
    "confirmed_catalog_year",
    "visual_evidence_text",
    "visual_confidence",
    "visual_notes",
    "local_pdf_path",
    "page_image_path",
    "api_call_id",
    "parsed_response_path",
    "created_at",
]


def read_optional_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, low_memory=False)


def read_strict_outputs(
    repo_root: Path,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
]:
    return (
        pd.read_csv(repo_root / STRICT_YEAR_COVERAGE_INPUT, low_memory=False),
        pd.read_csv(repo_root / STRICT_RETRIEVAL_COVERAGE_INPUT, low_memory=False),
        pd.read_csv(repo_root / STRICT_INVENTORY_INPUT, low_memory=False),
        read_optional_csv(repo_root / PANEL_YEAR_STATUS_INPUT),
        read_optional_csv(repo_root / PANEL_CANDIDATES_INPUT),
        read_optional_csv(repo_root / PANEL_RETRIEVAL_COVERAGE_INPUT),
        read_optional_csv(repo_root / PANEL_RETRIEVED_YEAR_COVERAGE_INPUT),
        read_optional_csv(repo_root / SOURCE_ROOT_PLAN_INPUT),
        read_optional_csv(repo_root / ESCALATION_QUEUE_INPUT),
        read_optional_csv(repo_root / CURRENT_SOURCE_TRACE_INPUT),
        read_optional_csv(repo_root / CURRENT_YEAR_TRACE_INPUT),
        read_optional_csv(repo_root / OCR_VISUAL_CONFIRMATION_INPUT),
    )


def build_summary(year_coverage: pd.DataFrame, retrieval_coverage: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        year_coverage.groupby(
            ["strict_pilot_rank", "unitid", "institution_name", "state", "strict_pilot_reason"],
            dropna=False,
        )
        .agg(
            institution_years=("target_year", "count"),
            covered_years=("has_strict_catalog_source", "sum"),
            legacy_evidence_years=("legacy_evidence_row_count", lambda values: int((values.fillna(0) > 0).sum())),
        )
        .reset_index()
    )
    grouped["covered_years"] = grouped["covered_years"].astype(int)
    grouped["missing_years"] = grouped["institution_years"] - grouped["covered_years"]
    grouped["coverage_rate"] = (grouped["covered_years"] / grouped["institution_years"]).round(3)

    source_counts = (
        retrieval_coverage.groupby(["unitid"], dropna=False)
        .agg(
            source_rows=("source_id", "count"),
            strict_source_rows=("strict_covers_target_year", "sum"),
            review_source_rows=("strict_covers_target_year", lambda values: int((~values.fillna(False)).sum())),
        )
        .reset_index()
    )
    out = grouped.merge(source_counts, on="unitid", how="left")
    for col in ["source_rows", "strict_source_rows", "review_source_rows"]:
        out[col] = out[col].fillna(0).astype(int)
    return out[SUMMARY_COLUMNS].sort_values(["strict_pilot_rank", "unitid"])


def build_needs_review(year_coverage: pd.DataFrame, retrieval_coverage: pd.DataFrame) -> pd.DataFrame:
    missing_years = year_coverage.loc[~year_coverage["has_strict_catalog_source"].fillna(False)].copy()
    missing_years["review_type"] = "missing_institution_year"
    missing_years["best_final_url"] = ""
    missing_years["legacy_review_reasons"] = ""

    source_review = retrieval_coverage.loc[~retrieval_coverage["strict_covers_target_year"].fillna(False)].copy()
    source_review["review_type"] = "source_not_strict_coverage"
    source_review["strict_pilot_rank"] = source_review["pilot_rank"]
    source_review["review_reason"] = source_review["strict_coverage_reason"].fillna(source_review["review_reason"])

    combined = pd.concat(
        [missing_years.reindex(columns=NEEDS_REVIEW_COLUMNS), source_review.reindex(columns=NEEDS_REVIEW_COLUMNS)],
        ignore_index=True,
    )
    return combined.sort_values(["strict_pilot_rank", "unitid", "target_year", "review_type", "source_id"])


def select_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    return out[columns]


def write_review_workbook(
    year_coverage: pd.DataFrame,
    retrieval_coverage: pd.DataFrame,
    inventory: pd.DataFrame,
    panel_year_status: pd.DataFrame,
    panel_candidates: pd.DataFrame,
    panel_retrieval: pd.DataFrame,
    panel_retrieved_year_coverage: pd.DataFrame,
    source_root_plan: pd.DataFrame,
    escalation_queue: pd.DataFrame,
    current_source_trace: pd.DataFrame,
    current_year_trace: pd.DataFrame,
    ocr_visual_confirmation: pd.DataFrame,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = build_summary(year_coverage, retrieval_coverage)
    needs_review = build_needs_review(year_coverage, retrieval_coverage)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        select_columns(year_coverage, YEAR_COVERAGE_COLUMNS).to_excel(writer, sheet_name="year_coverage", index=False)
        select_columns(retrieval_coverage, SOURCE_EVIDENCE_COLUMNS).to_excel(
            writer, sheet_name="source_evidence", index=False
        )
        needs_review.to_excel(writer, sheet_name="needs_review", index=False)
        if not panel_year_status.empty:
            select_columns(panel_year_status, PANEL_YEAR_STATUS_COLUMNS).to_excel(
                writer, sheet_name="panel_year_status", index=False
            )
        if not panel_candidates.empty:
            select_columns(panel_candidates, PANEL_CANDIDATE_COLUMNS).to_excel(
                writer, sheet_name="panel_candidates", index=False
            )
        if not panel_retrieval.empty:
            select_columns(panel_retrieval, PANEL_RETRIEVAL_COLUMNS).to_excel(
                writer, sheet_name="panel_retrieval", index=False
            )
        if not panel_retrieved_year_coverage.empty:
            select_columns(panel_retrieved_year_coverage, PANEL_RETRIEVED_YEAR_COLUMNS).to_excel(
                writer, sheet_name="panel_retrieved_years", index=False
            )
        if not source_root_plan.empty:
            select_columns(source_root_plan, SOURCE_ROOT_PLAN_COLUMNS).to_excel(
                writer, sheet_name="source_root_plan", index=False
            )
        if not escalation_queue.empty:
            select_columns(escalation_queue, ESCALATION_QUEUE_COLUMNS).to_excel(
                writer, sheet_name="escalation_queue", index=False
            )
        if not current_source_trace.empty:
            select_columns(current_source_trace, CURRENT_SOURCE_TRACE_COLUMNS).to_excel(
                writer, sheet_name="current_source_trace", index=False
            )
        if not current_year_trace.empty:
            select_columns(current_year_trace, CURRENT_YEAR_TRACE_COLUMNS).to_excel(
                writer, sheet_name="current_year_trace", index=False
            )
        if not ocr_visual_confirmation.empty:
            select_columns(ocr_visual_confirmation, OCR_VISUAL_COLUMNS).to_excel(
                writer, sheet_name="ocr_visual_abac", index=False
            )
        inventory.to_excel(writer, sheet_name="inventory_provenance", index=False)
        format_workbook(writer.book)


def format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")
            values = [str(cell.value) for cell in column_cells[:250] if cell.value is not None]
            width = max([len(header), *(len(value) for value in values)] or [len(header)])
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(max(width + 2, 10), 70)


def run_strict_review_workbook(repo_root: Path) -> Path:
    repo_root = repo_root.resolve()
    (
        year_coverage,
        retrieval_coverage,
        inventory,
        panel_year_status,
        panel_candidates,
        panel_retrieval,
        panel_retrieved_year_coverage,
        source_root_plan,
        escalation_queue,
        current_source_trace,
        current_year_trace,
        ocr_visual_confirmation,
    ) = read_strict_outputs(repo_root)
    output_path = (repo_root / STRICT_REVIEW_WORKBOOK_OUTPUT).resolve()
    write_review_workbook(
        year_coverage,
        retrieval_coverage,
        inventory,
        panel_year_status,
        panel_candidates,
        panel_retrieval,
        panel_retrieved_year_coverage,
        source_root_plan,
        escalation_queue,
        current_source_trace,
        current_year_trace,
        ocr_visual_confirmation,
        output_path,
    )
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build strict Phase 3 catalog pilot review workbook.")
    parser.add_argument("--root", type=Path, default=None)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    root = args.root.resolve() if args.root else repo_root_from_cwd()
    output_path = run_strict_review_workbook(root)
    print(f"review_workbook: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
