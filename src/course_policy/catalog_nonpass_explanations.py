"""Write plain-language explanations for non-pass catalog audit rows."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .ai_config import repo_root_from_cwd
from .catalog_spotcheck_audit import SPOTCHECK_AUDIT_OUTPUT
from .spotcheck_workbook import SPOTCHECK_WORKBOOK_OUTPUT, format_workbook


DATA_DIR = Path("../data_policy_pipeline")
LOG_DIR = DATA_DIR / "logs"

NONPASS_EXPLANATIONS_OUTPUT = LOG_DIR / "catalog_url_spotcheck_nonpass_explanations.md"

RECENT_EXPANSION_UNITIDS = (
    104151,
    110617,
    110714,
    115755,
    123572,
    126562,
    128771,
    129020,
    130493,
    130943,
)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def year_list(values: pd.Series) -> str:
    years = sorted({int(value) for value in values.dropna().astype(int)})
    return "; ".join(str(year) for year in years) if years else "none"


def explanation_for(group: pd.DataFrame, audit_row: pd.Series) -> str:
    status = clean_text(audit_row["audit_status"])
    manual_status = clean_text(audit_row.get("manual_status", "")).lower()
    evidence = clean_text(group["manual_search_evidence"].iloc[0])
    accepted = clean_text(audit_row.get("accepted_stop_reasons", ""))
    ocr = clean_text(audit_row.get("ocr_or_visual_review_issues", ""))
    issues = clean_text(audit_row.get("pipeline_fix_issues", ""))

    if status == "needs_pipeline_fix":
        return (
            f"The audit still sees a pipeline problem: {issues or 'unexplained missing or wrong-scope rows'}. "
            "This institution should stay out of review-ready outputs until the source-finding or parsing rule is fixed."
        )
    if status == "needs_ocr_or_visual_review":
        return (
            "Candidate URLs were found, but the files still need OCR or visual confirmation before policy extraction. "
            f"This is an OCR queue item, not a source-discovery failure. {ocr}".strip()
        )
    if "catalog_dead_end" in manual_status:
        return (
            "Catalog-first discovery dead-ended because the available leads are institution-wide policy pages or school-specific catalogs, "
            "not a university-wide undergraduate catalog panel."
        )
    if "scope_dead_end" in manual_status:
        return (
            "The available catalog lead is outside the undergraduate catalog scope for this project, so the catalog-first process stops at scope review."
        )
    if "archive_bounds_explain_missing_years" in accepted:
        return (
            "The reviewed source root has an observed coverage boundary. Years outside that boundary are recorded as archive-bound stops, "
            "while years inside the observed span use the found catalog URLs."
        )
    if "row_level_verified_source_gaps" in accepted:
        return (
            "The reviewed source root was checked and exposes most of the span, but specific target years remain visibly absent or blocked. "
            "Those rows are recorded as verified source gaps for a later recovery pass."
        )
    if evidence:
        return evidence
    return "The row is accepted as a bounded source stop or deferred non-discovery queue item."


def build_nonpass_table(audit: pd.DataFrame, workbook: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    nonpass = audit.loc[~audit["audit_status"].eq("pass_basic_checks")].copy()
    for _, audit_row in nonpass.sort_values(["audit_status", "institution_name"]).iterrows():
        unitid = int(audit_row["unitid"])
        group = workbook.loc[workbook["unitid"].eq(unitid)].sort_values("start_year")
        blank_years = group.loc[group["best_url"].map(clean_text).eq(""), "start_year"]
        rows.append(
            {
                "unitid": unitid,
                "institution_name": audit_row["institution_name"],
                "audit_status": audit_row["audit_status"],
                "best_url_year_count": int(audit_row["best_url_year_count"]),
                "blank_or_stop_years": year_list(blank_years),
                "accepted_stop_reasons": clean_text(audit_row.get("accepted_stop_reasons", "")),
                "pipeline_fix_issues": clean_text(audit_row.get("pipeline_fix_issues", "")),
                "explanation": explanation_for(group, audit_row),
            }
        )
    return pd.DataFrame(rows)


def build_recent_summary(audit: pd.DataFrame) -> pd.DataFrame:
    recent = audit.loc[audit["unitid"].isin(RECENT_EXPANSION_UNITIDS)].copy()
    keep = [
        "unitid",
        "institution_name",
        "audit_status",
        "best_url_year_count",
        "legacy_url_year_count",
        "missing_best_url_years",
        "accepted_stop_reasons",
        "pipeline_fix_issues",
        "warnings",
        "manual_best_root_url",
        "manual_coverage_start_year",
        "manual_coverage_end_year",
    ]
    return recent[[col for col in keep if col in recent.columns]].sort_values(["audit_status", "institution_name"])


def build_report(repo_root: Path) -> str:
    audit = pd.read_csv(repo_root / SPOTCHECK_AUDIT_OUTPUT, low_memory=False)
    workbook = pd.read_excel(repo_root / SPOTCHECK_WORKBOOK_OUTPUT, sheet_name="spotcheck_mockup")
    nonpass = build_nonpass_table(audit, workbook)
    needs_fix_count = int(audit["audit_status"].eq("needs_pipeline_fix").sum())

    lines = [
        "# Non-Pass Catalog Audit Explanations",
        "",
        f"Generated at: {utc_now()}",
        "",
        "This note translates non-pass audit statuses into plain-language process explanations. These are not requests for manual row troubleshooting.",
        "",
        "## Summary",
        "",
        f"- `needs_pipeline_fix`: {needs_fix_count} institutions.",
        "- Non-pass institutions are either OCR/visual review, accepted archive/source gaps, scope dead ends, or catalog-first dead ends.",
        "",
    ]
    for _, row in nonpass.iterrows():
        unitid = int(row["unitid"])
        lines.extend(
            [
                f"## {row['institution_name']} ({unitid})",
                "",
                f"- Audit status: `{row['audit_status']}`",
                f"- Best URL years filled: {int(row['best_url_year_count'])} of 21",
                f"- Blank/stop years: {row['blank_or_stop_years']}",
                f"- Explanation: {row['explanation']}",
                "",
            ]
        )
    return "\n".join(lines)


def write_workbook_navigation_sheets(repo_root: Path) -> None:
    workbook_path = repo_root / SPOTCHECK_WORKBOOK_OUTPUT
    audit = pd.read_csv(repo_root / SPOTCHECK_AUDIT_OUTPUT, low_memory=False)
    spotcheck = pd.read_excel(workbook_path, sheet_name="spotcheck_mockup")
    recent_rows = spotcheck.loc[spotcheck["unitid"].isin(RECENT_EXPANSION_UNITIDS)].copy()
    recent_summary = build_recent_summary(audit)
    nonpass = build_nonpass_table(audit, spotcheck)

    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        recent_summary.to_excel(writer, sheet_name="recent_additions_summary", index=False)
        recent_rows.to_excel(writer, sheet_name="recent_additions", index=False)
        nonpass.to_excel(writer, sheet_name="nonpass_explanations", index=False)

    wb = load_workbook(workbook_path)
    preferred_order = [
        "summary",
        "recent_additions_summary",
        "recent_additions",
        "nonpass_explanations",
        "spotcheck_mockup",
        "missing_best_url",
        "legacy_comparison",
    ]
    wb._sheets = [wb[name] for name in preferred_order if name in wb.sheetnames] + [
        sheet for sheet in wb.worksheets if sheet.title not in preferred_order
    ]
    wb.save(workbook_path)
    format_workbook(workbook_path)


def run(repo_root: Path) -> Path:
    repo_root = repo_root.resolve()
    output_path = repo_root / NONPASS_EXPLANATIONS_OUTPUT
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(build_report(repo_root) + "\n", encoding="utf-8")
    write_workbook_navigation_sheets(repo_root)
    return output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=None)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    repo_root = args.repo_root or repo_root_from_cwd()
    output_path = run(repo_root)
    print(f"nonpass_explanations: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
