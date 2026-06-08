"""Build Excel review mockups for catalog URL spot checks."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_config import repo_root_from_cwd
from .review_ready_adjustments import ACCEPTED_SOURCE_GAPS, SUPPLEMENTAL_CATALOG_CANDIDATES


DATA_DIR = Path("../data_policy_pipeline")
INTERIM_DIR = DATA_DIR / "interim"
REVIEW_DIR = DATA_DIR / "review"
LOG_DIR = DATA_DIR / "logs"

LEGACY_EVIDENCE_LINKS_INPUT = INTERIM_DIR / "legacy_evidence_links.csv"
INSTITUTION_YEAR_TARGETS_INPUT = INTERIM_DIR / "institution_year_targets.csv"
STAGE_STATUS_INPUTS = (
    INTERIM_DIR / "catalog_batch2_stage_status.csv",
    INTERIM_DIR / "catalog_batch3_stage_status.csv",
    INTERIM_DIR / "catalog_batch4_stage_status.csv",
)
MANUAL_CATALOG_SEARCH_AUDIT_INPUT = REVIEW_DIR / "manual_catalog_search_audit.csv"
REVIEWED_ROOT_YEAR_COVERAGE_INPUT = INTERIM_DIR / "catalog_reviewed_root_year_coverage.csv"

SPOTCHECK_WORKBOOK_OUTPUT = REVIEW_DIR / "catalog_url_spotcheck_mockup.xlsx"
SPOTCHECK_SUMMARY_OUTPUT = LOG_DIR / "catalog_url_spotcheck_mockup_summary.md"


MOCKUP_COLUMNS = [
    "unitid",
    "institution_name",
    "grade_averaging",
    "grade_avg_threshold",
    "grade_forgiveness",
    "grade_forgive_threshold",
    "start_year",
    "best_url",
    "legacy_url",
    "evidence_text",
    "comments",
    "best_url_source",
    "best_url_status",
    "legacy_url_count",
    "catalog_title_or_link_text",
    "pipeline_stage",
    "stop_reason",
    "next_batch_action",
    "preferred_source_root_url",
    "archive_url",
    "retrieval_status",
    "source_retrieved",
    "legacy_workbook",
    "legacy_excel_rows",
    "legacy_review_reasons",
    "legacy_excerpt",
    "stage_explanation",
    "review_note",
    "manual_status",
    "manual_best_root_url",
    "manual_root_type",
    "manual_coverage_start_year",
    "manual_coverage_end_year",
    "manual_search_evidence",
    "programmatic_fix_needed",
    "next_pipeline_action_from_manual",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def first_nonempty(values: pd.Series) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def joined_unique(values: pd.Series, *, limit: int = 3) -> str:
    out: list[str] = []
    for value in values:
        text = clean_text(value)
        if text and text not in out:
            out.append(text)
    return "; ".join(out[:limit])


def read_stage_status(repo_root: Path) -> pd.DataFrame:
    frames = []
    for path in STAGE_STATUS_INPUTS:
        full_path = repo_root / path
        if not full_path.exists():
            continue
        frame = pd.read_csv(full_path, low_memory=False)
        frame["discovery_batch"] = path.name.replace("catalog_", "").replace("_stage_status.csv", "")
        frames.append(frame)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def read_legacy_links(repo_root: Path) -> pd.DataFrame:
    path = repo_root / LEGACY_EVIDENCE_LINKS_INPUT
    legacy = pd.read_csv(path, low_memory=False)
    legacy["legacy_url"] = legacy["legacy_url"].map(clean_text)
    legacy = legacy.loc[legacy["legacy_workbook"].eq("public")].copy()
    grouped = (
        legacy.sort_values(["unitid", "target_year", "selected_as_prior_evidence", "legacy_source_priority"], ascending=[True, True, False, True])
        .groupby(["unitid", "target_year"], as_index=False)
        .agg(
            legacy_url=("legacy_url", joined_unique),
            legacy_url_count=("legacy_url", lambda values: int(values.map(clean_text).ne("").sum())),
            legacy_workbook=("legacy_workbook", joined_unique),
            legacy_excel_rows=("legacy_excel_row", lambda values: joined_unique(values.astype(str), limit=8)),
            grade_averaging=("grade_averaging", first_nonempty),
            grade_avg_threshold=("grade_avg_threshold", first_nonempty),
            grade_forgiveness=("grade_forgiveness", first_nonempty),
            grade_forgive_threshold=("grade_forgive_threshold", first_nonempty),
            legacy_policy_class=("legacy_policy_class", first_nonempty),
            legacy_review_reasons=("legacy_review_reasons", joined_unique),
            legacy_excerpt=("legacy_excerpt", first_nonempty),
        )
    )
    return grouped


def read_manual_catalog_search_audit(repo_root: Path) -> pd.DataFrame:
    path = repo_root / MANUAL_CATALOG_SEARCH_AUDIT_INPUT
    if not path.exists():
        return pd.DataFrame(columns=["unitid"])
    audit = pd.read_csv(path, low_memory=False)
    keep_cols = [
        "unitid",
        "manual_status",
        "manual_best_root_url",
        "manual_root_type",
        "manual_coverage_start_year",
        "manual_coverage_end_year",
        "manual_search_evidence",
        "programmatic_fix_needed",
        "next_pipeline_action",
    ]
    audit = audit[[col for col in keep_cols if col in audit.columns]].drop_duplicates("unitid")
    return audit.rename(columns={"next_pipeline_action": "next_pipeline_action_from_manual"})


def read_manual_targets(repo_root: Path, manual_audit: pd.DataFrame) -> pd.DataFrame:
    if manual_audit.empty:
        return pd.DataFrame()
    path = repo_root / INSTITUTION_YEAR_TARGETS_INPUT
    targets = pd.read_csv(path, low_memory=False)
    unitids = set(manual_audit["unitid"].astype(int))
    rows = targets.loc[targets["unitid"].isin(unitids) & targets["year"].between(2000, 2020)].copy()
    rows = rows.rename(columns={"year": "target_year"})
    rows["pipeline_stage"] = "reviewed_root_panel"
    rows["stop_reason"] = ""
    rows["next_batch_action"] = ""
    rows["stage_explanation"] = "Institution-year row added from reviewed source-root audit."
    return rows


def read_reviewed_root_year_coverage(repo_root: Path) -> pd.DataFrame:
    path = repo_root / REVIEWED_ROOT_YEAR_COVERAGE_INPUT
    if not path.exists():
        return pd.DataFrame(columns=["unitid", "start_year"])
    coverage = pd.read_csv(path, low_memory=False)
    coverage = coverage.rename(columns={"target_year": "start_year"})
    coverage["reviewed_candidate_url"] = coverage.get("candidate_url", "").map(clean_text)
    if coverage.empty:
        return pd.DataFrame(columns=["unitid", "start_year"])
    keep_cols = [
        "unitid",
        "start_year",
        "reviewed_candidate_url",
        "candidate_link_text",
        "candidate_evidence_text",
        "candidate_evidence_source",
        "candidate_source_method",
        "archive_url",
        "reviewed_root_url",
        "validation_status",
        "reviewed_candidate_status",
    ]
    coverage = coverage[[col for col in keep_cols if col in coverage.columns]].drop_duplicates(["unitid", "start_year"])
    return coverage.rename(
        columns={
            "candidate_link_text": "reviewed_candidate_link_text",
            "candidate_evidence_text": "reviewed_candidate_evidence_text",
            "candidate_evidence_source": "reviewed_candidate_evidence_source",
            "candidate_source_method": "reviewed_candidate_source_method",
            "archive_url": "reviewed_archive_url",
            "reviewed_root_url": "reviewed_source_root_url",
            "validation_status": "reviewed_validation_status",
        }
    )


def best_url_for_row(row: pd.Series) -> tuple[str, str, str, str]:
    candidate_url = clean_text(row.get("candidate_url", ""))
    retrieved_url = clean_text(row.get("retrieved_candidate_url", ""))
    policy_page_url = clean_text(row.get("legacy_policy_page_url", ""))
    legacy_url = clean_text(row.get("legacy_url", ""))
    candidate_title = clean_text(row.get("candidate_link_text", "")) or clean_text(row.get("retrieved_candidate_link_text", ""))

    if candidate_url:
        return candidate_url, "preferred_or_secondary_archive_candidate", clean_text(row.get("retrieval_status", "")), candidate_title
    if retrieved_url:
        return retrieved_url, clean_text(row.get("retrieved_candidate_method", "")) or "retrieved_fallback_candidate", clean_text(row.get("retrieval_status", "")), candidate_title
    if policy_page_url:
        return policy_page_url, "legacy_policy_page_deferred", "policy_dating_needed", clean_text(row.get("retrieved_candidate_link_text", ""))
    if legacy_url:
        return legacy_url, "legacy_url_only", "not_checked_in_current_stage", ""
    return "", "", "", ""


def build_mockup(repo_root: Path) -> pd.DataFrame:
    stage = read_stage_status(repo_root)
    legacy = read_legacy_links(repo_root)
    manual_audit = read_manual_catalog_search_audit(repo_root)
    manual_targets = read_manual_targets(repo_root, manual_audit)
    reviewed_coverage = read_reviewed_root_year_coverage(repo_root)
    if not manual_targets.empty:
        if stage.empty:
            stage = manual_targets
        else:
            stage = pd.concat([stage, manual_targets], ignore_index=True, sort=False)
            stage = stage.drop_duplicates(["unitid", "target_year"], keep="first")
    if stage.empty:
        rows = legacy.rename(columns={"target_year": "start_year"}).copy()
        rows["best_url"] = rows["legacy_url"]
        rows["best_url_source"] = "legacy_url_only"
        rows["best_url_status"] = "not_checked_in_current_stage"
        rows["catalog_title_or_link_text"] = ""
        rows["review_note"] = "No batch stage-status files were available."
    else:
        rows = stage.merge(legacy, on=["unitid", "target_year"], how="left")
        rows = rows.rename(columns={"target_year": "start_year"})
        best = rows.apply(best_url_for_row, axis=1, result_type="expand")
        best.columns = ["best_url", "best_url_source", "best_url_status", "catalog_title_or_link_text"]
        for column in best.columns:
            rows[column] = best[column]
        rows["review_note"] = ""
    for column in MOCKUP_COLUMNS:
        if column not in rows.columns:
            rows[column] = ""
    rows["evidence_text"] = rows["legacy_excerpt"].map(clean_text)
    rows["comments"] = rows["stage_explanation"].map(clean_text)
    if not reviewed_coverage.empty:
        rows = rows.merge(reviewed_coverage, on=["unitid", "start_year"], how="left")
        reviewed_root_mask = rows["reviewed_source_root_url"].fillna("").astype(str).str.strip().ne("")
        rows.loc[reviewed_root_mask, "preferred_source_root_url"] = rows.loc[
            reviewed_root_mask, "reviewed_source_root_url"
        ].fillna(rows.loc[reviewed_root_mask, "preferred_source_root_url"])
        rows.loc[
            reviewed_root_mask
            & rows["best_url"].fillna("").astype(str).str.strip().eq("")
            & rows["stop_reason"].fillna("").astype(str).str.strip().eq(""),
            "stop_reason",
        ] = "reviewed_root_no_year_candidate"
        rows.loc[
            reviewed_root_mask
            & rows["best_url"].fillna("").astype(str).str.strip().eq("")
            & rows["next_batch_action"].fillna("").astype(str).str.strip().eq(""),
            "next_batch_action",
        ] = "continue_reviewed_root_expansion"
        reviewed_mask = rows["reviewed_candidate_url"].fillna("").astype(str).str.strip().ne("")
        rows.loc[reviewed_mask, "best_url"] = rows.loc[reviewed_mask, "reviewed_candidate_url"]
        rows.loc[reviewed_mask, "best_url_source"] = rows.loc[reviewed_mask, "reviewed_candidate_source_method"].fillna(
            "reviewed_root_candidate"
        )
        rows.loc[reviewed_mask, "best_url_status"] = rows.loc[reviewed_mask, "reviewed_validation_status"].fillna(
            "reviewed_candidate_found"
        )
        rows.loc[reviewed_mask, "catalog_title_or_link_text"] = rows.loc[reviewed_mask, "reviewed_candidate_link_text"].fillna("")
        rows.loc[reviewed_mask, "archive_url"] = rows.loc[reviewed_mask, "reviewed_archive_url"].fillna(
            rows.loc[reviewed_mask, "archive_url"]
        )
        rows.loc[reviewed_mask, "preferred_source_root_url"] = rows.loc[reviewed_mask, "reviewed_source_root_url"].fillna(
            rows.loc[reviewed_mask, "preferred_source_root_url"]
        )
        rows.loc[reviewed_mask, "stop_reason"] = "reviewed_candidate_found"
        rows.loc[reviewed_mask, "next_batch_action"] = "retrieve_or_extract_candidate"
        rows.loc[reviewed_mask, "review_note"] = "Best URL updated from reviewed source-root expansion."
    if not manual_audit.empty:
        rows = rows.merge(manual_audit, on="unitid", how="left", suffixes=("", "_manual"))
        for column in [
            "manual_status",
            "manual_best_root_url",
            "manual_root_type",
            "manual_coverage_start_year",
            "manual_coverage_end_year",
            "manual_search_evidence",
            "programmatic_fix_needed",
            "next_pipeline_action_from_manual",
        ]:
            manual_column = f"{column}_manual"
            if manual_column in rows.columns:
                rows[column] = rows[column].fillna("").where(rows[column].fillna("").ne(""), rows[manual_column].fillna(""))
                rows = rows.drop(columns=[manual_column])
        manual_root_mask = (
            rows["manual_best_root_url"].fillna("").astype(str).str.strip().ne("")
            & ~rows["manual_status"].fillna("").astype(str).str.contains("scope_dead_end", case=False, na=False)
        )
        manual_start = pd.to_numeric(rows["manual_coverage_start_year"], errors="coerce").fillna(0)
        manual_end = pd.to_numeric(rows["manual_coverage_end_year"], errors="coerce").fillna(0)
        manual_span_mask = (
            manual_start.eq(0)
            | manual_end.eq(0)
            | rows["start_year"].astype(int).between(manual_start.astype(int), manual_end.astype(int))
        )
        manual_root_mask = manual_root_mask & manual_span_mask
        rows.loc[manual_root_mask, "preferred_source_root_url"] = rows.loc[manual_root_mask, "manual_best_root_url"]
        missing_manual_root_mask = manual_root_mask & rows["best_url"].fillna("").astype(str).str.strip().eq("")
        rows.loc[missing_manual_root_mask, "archive_url"] = rows.loc[missing_manual_root_mask, "manual_best_root_url"]
        rows.loc[missing_manual_root_mask, "stop_reason"] = "reviewed_root_gap_unfilled"
        rows.loc[missing_manual_root_mask, "next_batch_action"] = rows.loc[
            missing_manual_root_mask, "next_pipeline_action_from_manual"
        ].fillna("continue_reviewed_root_expansion")
        rows.loc[
            missing_manual_root_mask & rows["review_note"].fillna("").astype(str).str.strip().eq(""),
            "review_note",
        ] = "Reviewed source root is identified, but the current parser did not produce a year-level candidate."
        manual_root_available = (
            rows["manual_best_root_url"].fillna("").astype(str).str.strip().ne("")
            & ~rows["manual_status"].fillna("").astype(str).str.contains("scope_dead_end", case=False, na=False)
            & rows["best_url"].fillna("").astype(str).str.strip().eq("")
        )
        lower_bound_mask = manual_root_available & manual_start.gt(0) & rows["start_year"].astype(int).lt(manual_start.astype(int))
        upper_bound_mask = manual_root_available & manual_end.gt(0) & rows["start_year"].astype(int).gt(manual_end.astype(int))
        for mask, reason in [
            (lower_bound_mask, "official_archive_lower_bound_reached"),
            (upper_bound_mask, "official_archive_upper_bound_reached"),
        ]:
            rows.loc[mask, "preferred_source_root_url"] = rows.loc[mask, "manual_best_root_url"]
            rows.loc[mask, "archive_url"] = rows.loc[mask, "manual_best_root_url"]
            rows.loc[mask, "stop_reason"] = reason
            rows.loc[mask, "next_batch_action"] = "record_official_coverage_limit"
            rows.loc[mask, "review_note"] = "Reviewed source root has an observed coverage bound outside this target year."
        catalog_dead_end_mask = (
            rows["manual_status"].fillna("").astype(str).str.contains("catalog_dead_end", case=False, na=False)
            & rows["best_url"].fillna("").astype(str).str.strip().eq("")
        )
        rows.loc[catalog_dead_end_mask, "stop_reason"] = "catalog_dead_end_wrong_scope"
        rows.loc[catalog_dead_end_mask, "next_batch_action"] = "catalog_dead_end"
        rows.loc[catalog_dead_end_mask, "review_note"] = "Catalog-first discovery dead-ended; preserve non-catalog policy leads for a later phase."
    supplemental_by_key = {
        (candidate.unitid, candidate.target_year): candidate
        for candidate in SUPPLEMENTAL_CATALOG_CANDIDATES
    }
    accepted_gap_by_key = {
        (gap.unitid, gap.target_year): gap
        for gap in ACCEPTED_SOURCE_GAPS
    }
    for index, row in rows.iterrows():
        key = (int(row["unitid"]), int(row["start_year"]))
        candidate = supplemental_by_key.get(key)
        if candidate and clean_text(row.get("best_url", "")) == "":
            rows.at[index, "best_url"] = candidate.candidate_url
            rows.at[index, "best_url_source"] = "reviewed_supplemental_candidate"
            rows.at[index, "best_url_status"] = "reviewed_candidate_found"
            rows.at[index, "catalog_title_or_link_text"] = candidate.candidate_link_text
            rows.at[index, "stop_reason"] = "reviewed_candidate_found"
            rows.at[index, "next_batch_action"] = "retrieve_or_extract_candidate"
            rows.at[index, "review_note"] = candidate.source_note
        gap = accepted_gap_by_key.get(key)
        if gap and clean_text(rows.at[index, "best_url"]) == "":
            rows.at[index, "stop_reason"] = gap.stop_reason
            rows.at[index, "next_batch_action"] = gap.next_batch_action
            rows.at[index, "review_note"] = gap.review_note
    rows = rows[MOCKUP_COLUMNS].sort_values(["institution_name", "start_year", "legacy_url"]).reset_index(drop=True)
    return rows


def summary_rows(mockup: pd.DataFrame) -> pd.DataFrame:
    if mockup.empty:
        return pd.DataFrame([{"metric": "rows", "value": 0}])
    rows = [
        {"metric": "generated_at", "value": utc_now()},
        {"metric": "rows", "value": len(mockup)},
        {"metric": "institution_count", "value": mockup["unitid"].nunique()},
        {"metric": "rows_with_best_url", "value": int(mockup["best_url"].map(clean_text).ne("").sum())},
        {"metric": "rows_with_legacy_url", "value": int(mockup["legacy_url"].map(clean_text).ne("").sum())},
        {"metric": "interior_archive_gap_rows", "value": int(mockup["stop_reason"].eq("interior_archive_gap").sum())},
    ]
    for source, count in mockup["best_url_source"].fillna("").replace("", "missing").value_counts().items():
        rows.append({"metric": f"best_url_source::{source}", "value": int(count)})
    return pd.DataFrame(rows)


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in ws.columns:
            header = clean_text(column_cells[0].value)
            max_len = max(len(clean_text(cell.value)) for cell in column_cells[:200])
            width = min(max(max_len + 2, len(header) + 2, 10), 70)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    wb.save(path)


def write_workbook(repo_root: Path, mockup: pd.DataFrame) -> Path:
    output = repo_root / SPOTCHECK_WORKBOOK_OUTPUT
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_rows(mockup).to_excel(writer, sheet_name="summary", index=False)
        mockup.to_excel(writer, sheet_name="spotcheck_mockup", index=False)
        mockup.loc[mockup["best_url"].map(clean_text).eq("")].to_excel(writer, sheet_name="missing_best_url", index=False)
        mockup.loc[mockup["legacy_url"].map(clean_text).ne("")].to_excel(writer, sheet_name="legacy_comparison", index=False)
    format_workbook(output)
    return output


def write_summary(repo_root: Path, mockup: pd.DataFrame, workbook_path: Path) -> Path:
    output = repo_root / SPOTCHECK_SUMMARY_OUTPUT
    output.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Catalog URL Spot-Check Mockup",
        "",
        f"Generated at: {utc_now()}",
        "",
        f"Workbook: `{workbook_path}`",
        "",
        "## Summary",
        "",
    ]
    for row in summary_rows(mockup).to_dict("records"):
        lines.append(f"- {row['metric']}: {row['value']}")
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output


def run(repo_root: Path) -> tuple[Path, Path]:
    repo_root = repo_root.resolve()
    mockup = build_mockup(repo_root)
    workbook_path = write_workbook(repo_root, mockup)
    summary_path = write_summary(repo_root, mockup, workbook_path)
    return workbook_path, summary_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=None, help="Path to policy_pipeline repo root.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    repo_root = args.repo_root or repo_root_from_cwd()
    workbook_path, summary_path = run(repo_root)
    print(f"workbook: {workbook_path}")
    print(f"summary: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
